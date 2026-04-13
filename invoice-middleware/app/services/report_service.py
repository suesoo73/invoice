from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.db.session import db_cursor

_PDF_FONT_NAME = "HYSMyeongJo-Medium"
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="F6ECDF")
_ITEM_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FCF4E8")
_HEADER_FONT = Font(bold=True)
_NUMBER_FORMAT = "#,##0"
_UNIT_PRICE_FORMAT = "#,##0.###"

pdfmetrics.registerFont(UnicodeCIDFont(_PDF_FONT_NAME))


def _filters(company_id: str, date_from: str | None, date_to: str | None) -> tuple[str, list[object]]:
    filters = ["company_id = %s", "status = 'completed'", "deleted_at IS NULL"]
    params: list[object] = [company_id]
    if date_from:
        filters.append("issue_date >= %s")
        params.append(date_from)
    if date_to:
        filters.append("issue_date <= %s")
        params.append(date_to)
    return " AND ".join(filters), params


def _normalize_include_tax(include_tax: bool | str | int | None) -> bool:
    if isinstance(include_tax, bool):
        return include_tax
    if include_tax is None:
        return True
    if isinstance(include_tax, int):
        return include_tax != 0
    return str(include_tax).strip().lower() in {"1", "true", "y", "yes", "on"}


def _format_number(value: object, digits: int = 0) -> str:
    if value is None or value == "":
        return "-"
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if digits <= 0:
        return f"{round(num):,}"
    return f"{num:,.{digits}f}"


def _period_sql(period_type: str) -> tuple[str, str, str]:
    if period_type == "weekly":
        start_expr = "DATE_SUB(issue_date, INTERVAL WEEKDAY(issue_date) DAY)"
        return (
            start_expr,
            "DATE_FORMAT(DATE_SUB(issue_date, INTERVAL WEEKDAY(issue_date) DAY), '%Y-%m-%d')",
            "주별",
        )
    if period_type == "monthly":
        start_expr = "DATE_FORMAT(issue_date, '%Y-%m-01')"
        return (
            start_expr,
            "DATE_FORMAT(issue_date, '%Y-%m')",
            "월별",
        )
    if period_type == "quarterly":
        start_expr = "MAKEDATE(YEAR(issue_date), 1) + INTERVAL QUARTER(issue_date) QUARTER - INTERVAL 1 QUARTER"
        return (
            start_expr,
            "CONCAT(YEAR(issue_date), '-Q', QUARTER(issue_date))",
            "분기별",
        )
    if period_type == "yearly":
        start_expr = "MAKEDATE(YEAR(issue_date), 1)"
        return (
            start_expr,
            "CAST(YEAR(issue_date) AS CHAR)",
            "연도별",
        )
    raise ValueError("Unsupported period_type")


def _query_report_documents(
    *,
    company_id: str,
    date_from: str | None = None,
    date_to: str | None = None,
) -> list[dict]:
    where_clause, params = _filters(company_id, date_from, date_to)
    with db_cursor(dictionary=True) as (_, cursor):
        cursor.execute(
            f"""
            SELECT
                id,
                issue_date,
                vendor_name,
                supply_amount,
                tax_amount,
                total_amount,
                status,
                original_filename,
                created_at
            FROM documents
            WHERE {where_clause}
            ORDER BY issue_date DESC, created_at DESC, id DESC
            """,
            tuple(params),
        )
        documents = cursor.fetchall()

        document_ids = [row["id"] for row in documents]
        items_by_document: dict[str, list[dict]] = {}
        if document_ids:
            placeholders = ", ".join(["%s"] * len(document_ids))
            cursor.execute(
                f"""
                SELECT
                    document_id,
                    line_no,
                    item_name,
                    quantity,
                    unit_price,
                    line_amount,
                    tax_amount,
                    total_amount
                FROM document_items
                WHERE document_id IN ({placeholders})
                ORDER BY document_id, line_no, id
                """,
                tuple(document_ids),
            )
            for item in cursor.fetchall():
                items_by_document.setdefault(item["document_id"], []).append(item)

    for document in documents:
        document["items"] = items_by_document.get(document["id"], [])
    return documents


def _query_report_summary(
    *,
    company_id: str,
    period_type: str = "monthly",
    date_from: str | None = None,
    date_to: str | None = None,
    include_tax: bool | str | int | None = True,
) -> dict:
    where_clause, params = _filters(company_id, date_from, date_to)
    normalized_include_tax = _normalize_include_tax(include_tax)
    amount_expr = "total_amount" if normalized_include_tax else "supply_amount"
    item_amount_expr = "i.total_amount" if normalized_include_tax else "i.line_amount"
    period_sort_expr, period_label_expr, period_label_name = _period_sql(period_type)
    amount_basis_label = "합계금액" if normalized_include_tax else "공급가액"

    with db_cursor(dictionary=True) as (_, cursor):
        cursor.execute(
            f"""
            SELECT
                COUNT(*) AS document_count,
                IFNULL(SUM(supply_amount), 0) AS supply_amount_sum,
                IFNULL(SUM(tax_amount), 0) AS tax_amount_sum,
                IFNULL(SUM(total_amount), 0) AS total_amount_sum,
                IFNULL(SUM({amount_expr}), 0) AS purchase_amount_sum
            FROM documents
            WHERE {where_clause}
            """,
            tuple(params),
        )
        summary = cursor.fetchone()

        cursor.execute(
            f"""
            SELECT
                {period_label_expr} AS period,
                COUNT(*) AS document_count,
                IFNULL(SUM(supply_amount), 0) AS supply_amount_sum,
                IFNULL(SUM(tax_amount), 0) AS tax_amount_sum,
                IFNULL(SUM(total_amount), 0) AS total_amount_sum,
                IFNULL(SUM({amount_expr}), 0) AS purchase_amount_sum
            FROM documents
            WHERE {where_clause}
              AND issue_date IS NOT NULL
            GROUP BY {period_label_expr}
            ORDER BY MIN({period_sort_expr}) ASC
            """,
            tuple(params),
        )
        periods = cursor.fetchall()

        cursor.execute(
            f"""
            SELECT
                vendor_name,
                COUNT(*) AS document_count,
                IFNULL(SUM(supply_amount), 0) AS supply_amount_sum,
                IFNULL(SUM(tax_amount), 0) AS tax_amount_sum,
                IFNULL(SUM(total_amount), 0) AS total_amount_sum,
                IFNULL(SUM({amount_expr}), 0) AS purchase_amount_sum
            FROM documents
            WHERE {where_clause}
              AND vendor_name IS NOT NULL
            GROUP BY vendor_name
            ORDER BY purchase_amount_sum DESC, vendor_name ASC
            LIMIT 10
            """,
            tuple(params),
        )
        vendors = cursor.fetchall()

        cursor.execute(
            f"""
            SELECT
                i.item_name,
                COUNT(*) AS line_count,
                IFNULL(SUM(i.line_amount), 0) AS line_amount_sum,
                IFNULL(SUM(i.tax_amount), 0) AS tax_amount_sum,
                IFNULL(SUM(i.total_amount), 0) AS total_amount_sum,
                IFNULL(SUM({item_amount_expr}), 0) AS purchase_amount_sum
            FROM document_items i
            JOIN documents d ON d.id = i.document_id
            WHERE {where_clause.replace('company_id', 'd.company_id').replace('status', 'd.status').replace('deleted_at', 'd.deleted_at')}
              AND i.item_name IS NOT NULL
            GROUP BY i.item_name
            ORDER BY purchase_amount_sum DESC, i.item_name ASC
            LIMIT 10
            """,
            tuple(params),
        )
        items = cursor.fetchall()

    return {
        "summary": summary,
        "periods": periods,
        "vendors": vendors,
        "items": items,
        "filters": {
            "company_id": company_id,
            "period_type": period_type,
            "date_from": date_from,
            "date_to": date_to,
            "include_tax": normalized_include_tax,
            "period_label": period_label_name,
            "amount_basis_label": amount_basis_label,
        },
    }


def get_report_summary(
    *,
    company_id: str,
    period_type: str = "monthly",
    date_from: str | None = None,
    date_to: str | None = None,
    include_tax: bool | str | int | None = True,
) -> dict:
    return _query_report_summary(
        company_id=company_id,
        period_type=period_type,
        date_from=date_from,
        date_to=date_to,
        include_tax=include_tax,
    )


def _style_header_row(ws, row_index: int, fill: PatternFill = _HEADER_FILL) -> None:
    for cell in ws[row_index]:
        cell.font = _HEADER_FONT
        cell.fill = fill


def _set_column_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def build_report_xlsx(
    *,
    company_id: str,
    period_type: str = "monthly",
    date_from: str | None = None,
    date_to: str | None = None,
    include_tax: bool | str | int | None = True,
) -> BytesIO:
    report = _query_report_summary(
        company_id=company_id,
        period_type=period_type,
        date_from=date_from,
        date_to=date_to,
        include_tax=include_tax,
    )
    documents = _query_report_documents(
        company_id=company_id,
        date_from=date_from,
        date_to=date_to,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "문서관리"
    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = True

    ws.append(["날짜", "거래처(공급하는자)", "공급가액", "세액", "합계금액", "상태"])
    _style_header_row(ws, 1)
    _set_column_widths(ws, [14, 34, 14, 12, 14, 14])

    current_row = 2
    for document in documents:
        ws.append(
            [
                document["issue_date"],
                document["vendor_name"],
                document["supply_amount"],
                document["tax_amount"],
                document["total_amount"],
                str(document["status"] or "").upper(),
            ]
        )
        ws.cell(current_row, 3).number_format = _NUMBER_FORMAT
        ws.cell(current_row, 4).number_format = _NUMBER_FORMAT
        ws.cell(current_row, 5).number_format = _NUMBER_FORMAT
        detail_start = current_row + 1
        current_row += 1

        items = document.get("items") or []
        if items:
            ws.append(["", "품목명", "수량", "단가", "공급금액", "세액", "총금액"])
            _style_header_row(ws, current_row, _ITEM_HEADER_FILL)
            current_row += 1
            for item in items:
                ws.append(
                    [
                        "",
                        item["item_name"],
                        item["quantity"],
                        item["unit_price"],
                        item["line_amount"],
                        item["tax_amount"],
                        item["total_amount"],
                    ]
                )
                ws.cell(current_row, 3).number_format = _UNIT_PRICE_FORMAT
                ws.cell(current_row, 4).number_format = _UNIT_PRICE_FORMAT
                ws.cell(current_row, 5).number_format = _NUMBER_FORMAT
                ws.cell(current_row, 6).number_format = _NUMBER_FORMAT
                ws.cell(current_row, 7).number_format = _NUMBER_FORMAT
                current_row += 1
            ws.row_dimensions.group(detail_start, current_row - 1, outline_level=1, hidden=True)

    summary_ws = wb.create_sheet("요약")
    summary_ws.append(["항목", "값"])
    _style_header_row(summary_ws, 1)
    for key, value in report["summary"].items():
        summary_ws.append([key, value])

    periods_ws = wb.create_sheet("기간별")
    periods_ws.append(["기간", "문서 수", "공급가액", "세액", "합계금액", report["filters"]["amount_basis_label"]])
    _style_header_row(periods_ws, 1)
    for row in report["periods"]:
        periods_ws.append([row["period"], row["document_count"], row["supply_amount_sum"], row["tax_amount_sum"], row["total_amount_sum"], row["purchase_amount_sum"]])

    vendors_ws = wb.create_sheet("거래처별")
    vendors_ws.append(["거래처", "문서 수", "공급가액", "세액", "합계금액", report["filters"]["amount_basis_label"]])
    _style_header_row(vendors_ws, 1)
    for row in report["vendors"]:
        vendors_ws.append([row["vendor_name"], row["document_count"], row["supply_amount_sum"], row["tax_amount_sum"], row["total_amount_sum"], row["purchase_amount_sum"]])

    items_ws = wb.create_sheet("품목별")
    items_ws.append(["품목", "라인 수", "공급가액", "세액", "합계금액", report["filters"]["amount_basis_label"]])
    _style_header_row(items_ws, 1)
    for row in report["items"]:
        items_ws.append([row["item_name"], row["line_count"], row["line_amount_sum"], row["tax_amount_sum"], row["total_amount_sum"], row["purchase_amount_sum"]])

    for sheet in [summary_ws, periods_ws, vendors_ws, items_ws]:
        for row in sheet.iter_rows(min_row=2):
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = _NUMBER_FORMAT

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_report_pdf(
    *,
    company_id: str,
    period_type: str = "monthly",
    date_from: str | None = None,
    date_to: str | None = None,
    include_tax: bool | str | int | None = True,
) -> BytesIO:
    report = _query_report_summary(
        company_id=company_id,
        period_type=period_type,
        date_from=date_from,
        date_to=date_to,
        include_tax=include_tax,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    styles["Title"].fontName = _PDF_FONT_NAME
    styles["Title"].fontSize = 20
    styles["Title"].alignment = TA_CENTER
    styles["Heading2"].fontName = _PDF_FONT_NAME
    styles["Heading2"].fontSize = 14
    styles["Normal"].fontName = _PDF_FONT_NAME
    styles["Normal"].fontSize = 10

    amount_basis_label = report["filters"]["amount_basis_label"]
    story = [
        Paragraph("AI 매입 인사이트 리포트", styles["Title"]),
        Paragraph(f"기간 단위: {report['filters']['period_label']}", styles["Normal"]),
        Paragraph(f"조회 기간: {date_from or '-'} ~ {date_to or '-'}", styles["Normal"]),
        Paragraph(f"금액 기준: {amount_basis_label}", styles["Normal"]),
        Spacer(1, 10),
    ]

    def add_table(title: str, rows: list[list[object]]) -> None:
        story.append(Paragraph(title, styles["Heading2"]))
        table = Table(rows, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ead9c6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7c8b5")),
                    ("FONTNAME", (0, 0), (-1, 0), _PDF_FONT_NAME),
                    ("FONTNAME", (0, 1), (-1, -1), _PDF_FONT_NAME),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 10))

    add_table("요약", [["항목", "값"]] + [[key, _format_number(value)] for key, value in report["summary"].items()])
    add_table(
        "기간별",
        [["기간", "문서 수", "공급가액", "세액", "합계금액", amount_basis_label]]
        + [[row["period"], _format_number(row["document_count"]), _format_number(row["supply_amount_sum"]), _format_number(row["tax_amount_sum"]), _format_number(row["total_amount_sum"]), _format_number(row["purchase_amount_sum"])] for row in report["periods"]],
    )
    add_table(
        "거래처별",
        [["거래처", "문서 수", "공급가액", "세액", "합계금액", amount_basis_label]]
        + [[row["vendor_name"], _format_number(row["document_count"]), _format_number(row["supply_amount_sum"]), _format_number(row["tax_amount_sum"]), _format_number(row["total_amount_sum"]), _format_number(row["purchase_amount_sum"])] for row in report["vendors"]],
    )
    add_table(
        "품목별",
        [["품목", "라인 수", "공급가액", "세액", "합계금액", amount_basis_label]]
        + [[row["item_name"], _format_number(row["line_count"]), _format_number(row["line_amount_sum"]), _format_number(row["tax_amount_sum"]), _format_number(row["total_amount_sum"]), _format_number(row["purchase_amount_sum"])] for row in report["items"]],
    )

    doc.build(story)
    buffer.seek(0)
    return buffer
